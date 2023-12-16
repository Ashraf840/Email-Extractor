from pathlib import Path
import openpyxl
import datetime as dt
from utilities import EmailExtractor
from check_last_read_mail_date import latest_mail_reading_date
import pandas as pd
from timeit import default_timer as timer


class StoreDailyEmailToExcel:
    def __init__(self):
        self.email_extractor = EmailExtractor("INBOX")
        self.email_address = self.email_extractor.email_address
        self.email_address_normalized = "".join(ch for ch in self.email_address if ch.isalnum())
        self.emails_list = self.email_extractor.extract_emails(latest_mail_reading_date())
        self.target_file = Path(f"./{self.email_address_normalized}.xlsx")
        # self.target_file = Path("./xyz_excel.xlsx")

    def excelFileChecker(self):
        print("Target file existence:", self.target_file.is_file())
        return self.target_file.is_file()
    
    def createExcelFile(self):
        wb = openpyxl.Workbook()

        sheet = wb.create_sheet(index=0, title="Gmail")
        if 'Sheet' in wb.sheetnames:
                # print('Sheet exists')
                del wb['Sheet']

        # TODO: Use a list to call the values using list[0], list[1],...., etc while making the following lines into a single line.
        column_values = ["Email Uid", "Date", "From", "Subject", "Content"]
        sheet["A1"], sheet["B1"], sheet["C1"] = column_values[0], column_values[1], column_values[2] 
        sheet["D1"], sheet["E1"] = column_values[3], column_values[4]

        wb.save(self.target_file)

    def modifyExcelFile(self):
        if len(self.emails_list) > 0:
            print("Email list length:", len(self.emails_list))
            # print("Use a for loop to add each email into the row of excel file!")
            existing_dataset_file = f'./{self.email_address_normalized}.xlsx'
            existing_df = pd.read_excel(existing_dataset_file, engine='openpyxl')   # Omit the sheet_param, since the sheet_name renamed as "Sheet1"
            data_list = list()
            for email in self.emails_list:
                    UID, Date, From, Subject, Content = int(email['UID']), email['Date'], email['From'], email['Subject'], email['Content']
                    new_data = {
                        'Email Uid': UID,
                        'Date': Date,
                        'From': From,
                        'Subject': Subject,
                        'Content': Content,
                    }
                    data_list.append(new_data)

            new_df = pd.DataFrame(data_list)
            concatenated_df = pd.concat([existing_df, new_df], ignore_index=True)
            concatenated_df.to_excel(existing_dataset_file, index=False)
        
    
    def storeMailToExcelFile(self):
        if self.excelFileChecker():
            print("Only store mails to the excel file!")
            self.modifyExcelFile()
        else:
            print("Create excel file! Then store mails to that file!")
            self.createExcelFile()
            self.modifyExcelFile()


start = timer()
stdee = StoreDailyEmailToExcel()
stdee.storeMailToExcelFile()
end = timer()
print(f"Execution time: {(end - start)} seconds")