"""
1. Check if there is any excel file in the current dir. Otherwise, Create one using the required columns.
2. If a new excel needs to be created, then write the new record of row(s) after creation (create the columns firstly after creating the excel file).
3. Else, append the list of latest extracted email(s) at the last of the excel file.
"""

from pathlib import Path
import openpyxl
import datetime as dt
from utilities import get_environment_variables, EmailExtractor
from check_last_read_mail_date import latest_mail_reading_date
import pandas as pd
from timeit import default_timer as timer


start = timer()
email_address, password, imap_url = get_environment_variables()
# print("Email Address:", email_address)
email_address="".join(ch for ch in email_address if ch.isalnum())       # Normalize email address string
# print("Email Address - Normalized:", email_address)


target_file = Path(f"./{email_address}.xlsx")

if target_file.is_file():
      print("Excel file is present")

      ee = EmailExtractor("INBOX")
      emails_list = ee.extract_emails(latest_mail_reading_date())
      print("Email list length:", len(emails_list))
      if len(emails_list) > 0:
            # print("Use a for loop to add each email into the row of excel file!")
            existing_dataset_file = f'./{email_address}.xlsx'
            existing_df = pd.read_excel(existing_dataset_file, engine='openpyxl')   # Omit the sheet_param, since the sheet_name renamed as "Sheet1"
            # TODO: Check the length of the exisiting dataframes length. Because a threshold will be set for the excel file, a new excel file (sequel of the prev excel file) will be created to store newly fetched emails.
            # Extended TODO: The threshold would be determined as the sheets of the excel file. Both how many sheets can be hold, also based on the number of rows stored in each sheet. Normally there will be 12 sheets based on the 12 months of the year.
            data_list = list()
            for email in emails_list:
                  UID, Date, From, Subject, Content = int(email['UID']), email['Date'], email['From'], email['Subject'], email['Content']

                  new_data = {
                        'Email Uid': UID,
                        'Date': Date,
                        'From': From,
                        'Subject': Subject,
                        'Content': Content,
                  }
                  data_list.append(new_data)

            new_df = pd.DataFrame(data_list)
            concatenated_df = pd.concat([existing_df, new_df], ignore_index=True)   # What is "ignore_index=True" param?
            concatenated_df.to_excel(existing_dataset_file, index=False)
else:
      print("Excel file is not present")
      # filepath = f"/home/robin/Documents/Tanjim/Miscellaneous Projects/gmail_reader/{email_address}.xlsx"
      wb = openpyxl.Workbook()

      sheet = wb.create_sheet(index=0, title="Gmail")
      if 'Sheet' in wb.sheetnames:
            # print('Sheet exists')
            del wb['Sheet']

      sheet["A1"] = "Email Uid"
      sheet["B1"] = "Date"
      sheet["C1"] = "From"
      sheet["D1"] = "Subject"
      sheet["E1"] = "Content"

      wb.save(target_file)


      ee = EmailExtractor("INBOX")
      emails_list = ee.extract_emails(latest_mail_reading_date())
      print("Email list length:", len(emails_list))
      if len(emails_list) > 0:
            # print("Use a for loop to add each email into the row of excel file!")
            existing_dataset_file = f'./{email_address}.xlsx'
            existing_df = pd.read_excel(existing_dataset_file, engine='openpyxl')   # Omit the sheet_param, since the sheet_name renamed as "Sheet1"
            data_list = list()
            for email in emails_list:
                  UID, Date, From, Subject, Content = int(email['UID']), email['Date'], email['From'], email['Subject'], email['Content']
                  new_data = {
                        'Email Uid': UID,
                        'Date': Date,
                        'From': From,
                        'Subject': Subject,
                        'Content': Content,
                  }
                  data_list.append(new_data)

            new_df = pd.DataFrame(data_list)
            concatenated_df = pd.concat([existing_df, new_df], ignore_index=True)
            concatenated_df.to_excel(existing_dataset_file, index=False)
end = timer()
print(f"Execution time: {2*(end - start)} seconds")


